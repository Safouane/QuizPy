import os
import uuid
import json
import datetime
import random
import string
import copy
import sys  # Make sure sys is imported if used by get_base_dir implicitly
from pathlib import Path
from django.http import JsonResponse, HttpResponse  # Removed Http404 if not used

# from django.conf import settings # Not needed if using get_media_dir
from django.core.files.uploadedfile import UploadedFile
from authentication.decorators import api_teacher_required

# --- CORRECTED IMPORT ---
from core.json_storage import (
    load_data,
    save_data,
    get_media_dir,
)  # Import get_media_dir

# --- END CORRECTION ---
import traceback  # Import get_media_dir


from django.shortcuts import render
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import (
    csrf_exempt,
)  # Keep using exempt for now, manage CSRF properly with frontend later
from decimal import Decimal, ROUND_HALF_UP  # For accurate score calculation
import openpyxl  # Add import
from openpyxl.utils import get_column_letter  # Optional: For setting column widths


# Import the decorator we created in AUTH-3
from authentication.decorators import api_teacher_required


# --- Media File Handling Setup ---
MEDIA_ROOT_DIR = get_media_dir()  # <<< Use the correct helper function
MEDIA_URL_ROOT = "/media_files/"
if not MEDIA_ROOT_DIR:
    print(
        "CRITICAL ERROR [Views]: Media directory path could not be determined from json_storage."
    )
# --- End Media File Handling Setup ---


def save_uploaded_media(uploaded_file: UploadedFile) -> str | None:
    """Saves uploaded file to media dir with unique name, returns filename."""
    if not uploaded_file or not MEDIA_ROOT_DIR:
        return None
    try:
        max_size = 5 * 1024 * 1024  # 5MB limit
        if uploaded_file.size > max_size:
            print(f"WARNING: Uploaded file rejected (too large): {uploaded_file.name}")
            return None
        original_name = Path(uploaded_file.name)
        ext = original_name.suffix.lower()
        allowed_extensions = [".png", ".jpg", ".jpeg", ".gif", ".webp"]
        if ext not in allowed_extensions:
            print(
                f"WARNING: Uploaded file rejected (invalid type): {uploaded_file.name}"
            )
            return None
        filename = f"{uuid.uuid4()}{ext}"
        media_path = MEDIA_ROOT_DIR / filename
        print(f"DEBUG: Saving uploaded file '{uploaded_file.name}' as '{filename}'")
        with open(media_path, "wb+") as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)
        print(f"DEBUG: Successfully saved to: {media_path}")
        return filename
    except Exception as e:
        print(f"ERROR saving uploaded file {uploaded_file.name}: {e}")
        traceback.print_exc()
        return None


# --- End save_uploaded_media ---


# We will use these functions later to interact with CORE-3 JSON storage
# from core.json_storage import load_quiz_data, save_quiz_data
@csrf_exempt
@api_teacher_required
@require_http_methods(["GET", "POST"])
def quiz_list_create_api(request):
    """
    API endpoint for listing all quizzes (GET) or creating a new quiz (POST).
    Requires teacher authentication. Uses JSON storage.
    """
    data = load_data()  # Load current data from JSON
    quizzes = data.get("quizzes", [])

    if request.method == "GET":
        # Prepare data for response, ensuring no sets are included
        response_quizzes = []
        for quiz_data in quizzes:
            # Create a copy to modify without affecting original data if needed
            prepared_quiz = quiz_data.copy()
            # Explicitly convert potential sets within the quiz to lists
            if isinstance(prepared_quiz.get("questions"), set):
                prepared_quiz["questions"] = list(prepared_quiz["questions"])
            # Check other potential set fields within config if applicable
            # Example: if 'allowed_groups' in prepared_quiz.get('config', {}) and isinstance(prepared_quiz['config']['allowed_groups'], set):
            #     prepared_quiz['config']['allowed_groups'] = list(prepared_quiz['config']['allowed_groups'])
            response_quizzes.append(prepared_quiz)

        print(f"DEBUG: Returning quiz list. Count: {len(response_quizzes)}")
        # Return the cleaned list
        return JsonResponse({"quizzes": response_quizzes})

    elif request.method == "POST":
        try:
            import random  # Add import
            import string  # Add import

            # Get data from request body
            request_data = json.loads(request.body)
            title = request_data.get("title")
            description = request_data.get("description", "")  # Optional field

            if not title:
                return JsonResponse({"error": "Quiz title is required."}, status=400)

            def generate_quiz_key(length=6):
                # Simple key generator (letters and digits) - make more robust if needed
                characters = string.ascii_uppercase + string.digits
                return "".join(random.choice(characters) for i in range(length))

            # Generate a unique key (add checks later if collision is a concern)
            new_quiz_key = generate_quiz_key()

            # Create new quiz object
            new_quiz_id = str(uuid.uuid4())  # Generate a unique ID
            new_quiz = {
                "id": new_quiz_id,
                "title": title,
                "description": description,
                # Accept questions list, default to empty if not provided
                "questions": request_data.get("questions", []),
                "config": {  # Default config, can be updated later via PUT
                    "duration": None,  # in minutes
                    "pass_score": 70,  # percentage
                    "presentation_mode": "all",  # 'all' or 'one-by-one'
                    "allow_back": True,
                    "randomize_questions": False,
                    "shuffle_answers": False,
                },
                "access_key": new_quiz_key,
                "archived": False,
                "versions": [],  # For TCHR-8 later
                # Add created_at, updated_at timestamps if desired
            }

            # Add to the list and save
            quizzes.append(new_quiz)
            data["quizzes"] = quizzes  # Update the main data dict
            save_data(data)  # Write back to JSON file

            return JsonResponse(
                {"message": "Quiz created successfully.", "quiz": new_quiz}, status=201
            )

        except json.JSONDecodeError:
            return JsonResponse(
                {"error": "Invalid JSON format in request body."}, status=400
            )
        except Exception as e:
            # Log the exception e
            print(f"Error creating quiz: {e}")  # Basic logging
            return JsonResponse(
                {"error": f"An unexpected error occurred: {str(e)}"}, status=500
            )


# src/quiz/views.py


@csrf_exempt  # File uploads might need exemption or specific CSRF handling via JS
@api_teacher_required
@require_http_methods(["GET", "POST"])
def question_list_create_api(request):
    """
    API endpoint for listing questions (GET) or creating a new question (POST).
    POST handles multipart/form-data for potential file uploads.
    GET returns basic list (update later to include media filenames if needed by lists).
    """
    # --- GET Request Logic ---
    if request.method == "GET":
        print("DEBUG [GET Questions]: Fetching question list.")
        try:
            all_data = load_data()
            all_questions = all_data.get("questions", [])
            # Apply filters (example for category) - TODO: Refactor filtering logic
            category_filter = request.GET.get("category")
            type_filter = request.GET.get("type")  # Add other filters as needed

            filtered_questions = all_questions
            if category_filter:
                filtered_questions = [
                    q
                    for q in filtered_questions
                    if category_filter.lower() == q.get("category", "").lower()
                ]
            if type_filter:
                filtered_questions = [
                    q
                    for q in filtered_questions
                    if type_filter.upper() == q.get("type", "").upper()
                ]

            # Prepare simplified list for bank view / modal (can enhance later)
            prepared_questions = [
                {
                    "id": q.get("id"),
                    "text": q.get("text"),
                    "type": q.get("type"),
                    "category": q.get("category"),
                    "difficulty": q.get("difficulty"),
                    # Add media_filename thumbnail indicator later?
                }
                for q in filtered_questions
            ]
            return JsonResponse({"questions": prepared_questions})
        except Exception as e:
            print(f"ERROR [GET Questions]: {e}")
            traceback.print_exc()
            return JsonResponse({"error": "Failed to load questions"}, status=500)

    # --- POST Request Logic (Handles Form Data + Files) ---
    elif request.method == "POST":
        print("DEBUG [POST Question]: Received create request.")
        try:
            request_data = request.POST  # Use POST for form fields
            uploaded_files = request.FILES  # Use FILES for uploaded files
            print(f"DEBUG [POST Question]: Form data received: {request_data.dict()}")
            print(f"DEBUG [POST Question]: Files received: {uploaded_files.keys()}")

            # Extract Core Fields
            text = request_data.get("text", "").strip()
            q_type = request_data.get("type")
            score_str = request_data.get("score", "1")
            difficulty = request_data.get("difficulty", "Medium")
            category = request_data.get("category", "Uncategorized").strip()
            quiz_ids = request_data.getlist(
                "quiz_ids[]", []
            )  # Assumes sent as list if needed
            is_single_choice_mcq = (
                request_data.get("mcq_is_single_choice") == "true"
            )  # Check flag

            # Basic validation
            if not q_type:
                return JsonResponse({"error": "Question type is required."}, status=400)
            if (
                not text
                and not uploaded_files.get("question_media")
                and q_type != "MEDIA"
            ):  # Require text unless media provided
                return JsonResponse(
                    {"error": "Question text or media is required."}, status=400
                )
            try:  # Validate score format
                score = int(score_str) if score_str else 1
            except ValueError:
                return JsonResponse({"error": "Invalid score format."}, status=400)

            # Handle Question Media Upload
            question_media_file = uploaded_files.get("question_media")
            saved_question_media_filename = save_uploaded_media(question_media_file)

            # Prepare Base New Question Object
            new_question_id = str(uuid.uuid4())
            new_question = {
                "id": new_question_id,
                "quiz_ids": quiz_ids,
                "text": text,
                "type": q_type,
                "media_filename": saved_question_media_filename,
                "score": score,
                "difficulty": difficulty,
                "category": category,
                "options": [],
                "correct_answer": [],
                "short_answer_review_mode": "manual",
                "short_answer_correct_text": None,
                "mcq_is_single_choice": False,
            }

            # Handle MCQ Options & Media
            if q_type == "MCQ":
                new_question["mcq_is_single_choice"] = is_single_choice_mcq
                options_list = []
                correct_answer_ids = []
                i = 0
                while True:
                    # Check for indexed fields sent by frontend form
                    option_text_key = f"options_text[{i}]"
                    option_media_key = f"options_media[{i}]"
                    option_correct_key = f"options_correct[{i}]"  # Assume value 'on' or 'true' if checked

                    # Break loop if essential keys are missing for this index
                    if (
                        option_text_key not in request_data
                        and option_media_key not in uploaded_files
                    ):
                        print(
                            f"DEBUG [POST Question]: No more options found at index {i}."
                        )
                        break

                    option_text = request_data.get(option_text_key, "").strip()
                    option_media_file = uploaded_files.get(option_media_key)
                    is_correct_flag = request_data.get(option_correct_key) == "on"

                    saved_option_media_filename = save_uploaded_media(option_media_file)

                    # Require text OR media for an option
                    if not option_text and not saved_option_media_filename:
                        print(
                            f"WARNING [POST Question]: Skipping completely empty option at index {i}."
                        )
                        i += 1
                        continue

                    option_id = str(uuid.uuid4())
                    options_list.append(
                        {
                            "id": option_id,
                            "text": option_text if option_text else None,
                            "media_filename": saved_option_media_filename,
                        }
                    )
                    if is_correct_flag:
                        correct_answer_ids.append(option_id)
                    print(
                        f"DEBUG [POST Question]: Processed option {i}: Text='{option_text}', Media='{saved_option_media_filename}', Correct={is_correct_flag}"
                    )
                    i += 1
                # End Option Loop

                # Validate option/answer counts
                if len(options_list) < 2:
                    return JsonResponse(
                        {"error": "MCQ requires at least two valid options."},
                        status=400,
                    )
                if is_single_choice_mcq and len(correct_answer_ids) != 1:
                    return JsonResponse(
                        {
                            "error": "Single-choice MCQ requires exactly one correct answer."
                        },
                        status=400,
                    )
                if not is_single_choice_mcq and len(correct_answer_ids) == 0:
                    return JsonResponse(
                        {
                            "error": "Multiple-choice MCQ requires at least one correct answer."
                        },
                        status=400,
                    )

                new_question["options"] = options_list
                new_question["correct_answer"] = correct_answer_ids

            elif q_type == "SHORT_TEXT":
                # ... existing logic to process review mode and correct text ...
                review_mode = request_data.get("short_answer_review_mode", "manual")
                correct_text_raw = request_data.get("short_answer_correct_text")
                if review_mode not in ["manual", "auto"]:
                    return JsonResponse(
                        {"error": "Invalid short answer review mode."}, status=400
                    )
                if review_mode == "auto" and not correct_text_raw:
                    return JsonResponse(
                        {
                            "error": "Correct answer text is required for automatic review mode."
                        },
                        status=400,
                    )
                new_question["short_answer_review_mode"] = review_mode
                new_question["short_answer_correct_text"] = (
                    correct_text_raw.strip()
                    if review_mode == "auto" and correct_text_raw
                    else None
                )

            # Add Question and Save
            print(f"DEBUG [POST Question]: Final new question object: {new_question}")
            all_data = load_data()
            all_data.setdefault("questions", []).append(new_question)
            # TODO: Link question to quizzes specified in quiz_ids by modifying quiz objects in all_data
            save_data(all_data)
            print(f"DEBUG [POST Question]: Question {new_question_id} saved.")
            return JsonResponse(
                {"message": "Question created successfully.", "question": new_question},
                status=201,
            )

        except Exception as e:
            print(f"ERROR [POST Question]: Unexpected error: {e}")
            traceback.print_exc()
            return JsonResponse(
                {"error": f"An unexpected error occurred: {str(e)}"}, status=500
            )

    else:  # Should not happen due to require_http_methods
        return JsonResponse({"error": "Method not allowed"}, status=405)


# --- End question_list_create_api ---


@csrf_exempt
@api_teacher_required
@require_http_methods(["GET", "PUT", "DELETE"])
def question_detail_api(request, question_id):
    """
    API endpoint for retrieving (GET), updating (PUT - text/config ONLY),
    or deleting (DELETE) a specific question by its ID.
    """
    question_id_str = str(question_id)
    print(
        f"DEBUG [Q Detail API]: Request for Q ID: {question_id_str}, Method: {request.method}"
    )

    # --- GET Request Logic (Includes Media Filenames) ---
    if request.method == "GET":
        try:
            data = load_data()
            questions = data.get("questions", [])
            question = next(
                (q for q in questions if str(q.get("id")) == question_id_str), None
            )

            if question is None:
                return JsonResponse(
                    {"error": f"Question with ID {question_id_str} not found."},
                    status=404,
                )

            # Prepare response data - ensure media filenames are included
            prepared_question = copy.deepcopy(question)
            # Ensure options also include media_filename if they exist
            if prepared_question.get("type") == "MCQ" and isinstance(
                prepared_question.get("options"), list
            ):
                prepared_question["options"] = [
                    {
                        "id": opt.get("id"),
                        "text": opt.get("text"),
                        "media_filename": opt.get(
                            "media_filename"
                        ),  # <<< Include media filename
                    }
                    for opt in prepared_question["options"]
                    if isinstance(opt, dict)
                ]
            # Clean potential sets just in case (shouldn't occur with proper saving)
            if isinstance(prepared_question.get("correct_answer"), set):
                prepared_question["correct_answer"] = list(
                    prepared_question["correct_answer"]
                )
            if isinstance(prepared_question.get("quiz_ids"), set):
                prepared_question["quiz_ids"] = list(prepared_question["quiz_ids"])

            print(f"DEBUG [GET Q Detail]: Returning data: {prepared_question}")
            return JsonResponse({"question": prepared_question})
        except Exception as e:
            print(f"ERROR [GET Q Detail]: {e}")
            traceback.print_exc()
            return JsonResponse(
                {"error": "Failed to retrieve question details."}, status=500
            )

    # --- PUT Request Logic (Simplified - No Media Updates Here) ---
    elif request.method == "PUT":
        print(f"DEBUG [PUT Q Detail]: Request for Q ID: {question_id_str}")
        # --- IMPORTANT LIMITATION: Assumes JSON payload for text/config updates ONLY ---
        try:
            request_data = json.loads(request.body)  # Expect JSON payload
            print(f"DEBUG [PUT Q Detail]: Received JSON payload: {request_data}")

            data = load_data()
            questions = data.get("questions", [])
            question_to_update = None
            question_index = -1
            for index, q in enumerate(questions):
                if str(q.get("id")) == question_id_str:
                    question_to_update = q
                    question_index = index
                    break
            if question_to_update is None:
                return JsonResponse({"error": "Question not found."}, status=404)

            updated_question = copy.deepcopy(question_to_update)  # Work on a copy

            # --- Update ONLY non-file fields from JSON payload ---
            if "text" in request_data:
                updated_question["text"] = request_data["text"]
            if "score" in request_data:
                try:
                    updated_question["score"] = int(request_data["score"])
                except (ValueError, TypeError):
                    return JsonResponse({"error": "Invalid score format."}, status=400)
            if "difficulty" in request_data:
                updated_question["difficulty"] = request_data["difficulty"]
            if "category" in request_data:
                updated_question["category"] = request_data["category"]
            if "quiz_ids" in request_data:
                if not isinstance(request_data["quiz_ids"], list):
                    return JsonResponse(
                        {"error": "quiz_ids must be a list."}, status=400
                    )
                try:
                    updated_question["quiz_ids"] = [
                        str(qid) for qid in request_data["quiz_ids"]
                    ]
                except:
                    return JsonResponse(
                        {"error": "Invalid quiz_ids format."}, status=400
                    )

            # Update flags/modes
            if "mcq_is_single_choice" in request_data:
                updated_question["mcq_is_single_choice"] = bool(
                    request_data["mcq_is_single_choice"]
                )
            if "short_answer_review_mode" in request_data:
                mode = request_data["short_answer_review_mode"]
                if mode not in ["manual", "auto"]:
                    return JsonResponse({"error": "Invalid review mode."}, status=400)
                updated_question["short_answer_review_mode"] = mode
                if mode == "manual":
                    updated_question["short_answer_correct_text"] = (
                        None  # Clear correct text if switching to manual
                    )

            if "short_answer_correct_text" in request_data:
                if updated_question.get("short_answer_review_mode") == "auto":
                    text = request_data["short_answer_correct_text"]
                    updated_question["short_answer_correct_text"] = (
                        text.strip() if isinstance(text, str) else None
                    )
                else:
                    updated_question["short_answer_correct_text"] = (
                        None  # Ignore if not in auto mode
                    )

            # Update MCQ options TEXT ONLY and correct answers based on TEXTS
            if updated_question.get("type") == "MCQ":
                is_single = updated_question.get("mcq_is_single_choice", False)
                # Only update if relevant keys present
                if "options" in request_data or "correct_answer_texts" in request_data:
                    # Use existing options as base to preserve IDs and media
                    existing_options_map_by_text = {
                        opt.get("text"): opt
                        for opt in updated_question.get("options", [])
                        if isinstance(opt, dict) and opt.get("text")
                    }
                    options_texts = request_data.get(
                        "options"
                    )  # Expect list of strings
                    correct_texts = request_data.get(
                        "correct_answer_texts"
                    )  # Expect list of strings

                    if options_texts is None or correct_texts is None:
                        return JsonResponse(
                            {
                                "error": 'For MCQ update via JSON, provide lists for "options" (texts) and "correct_answer_texts".'
                            },
                            status=400,
                        )
                    if not isinstance(options_texts, list) or not isinstance(
                        correct_texts, list
                    ):
                        return JsonResponse(
                            {
                                "error": '"options" and "correct_answer_texts" must be lists.'
                            },
                            status=400,
                        )

                    # Validation
                    if is_single and len(correct_texts) != 1:
                        return JsonResponse(
                            {
                                "error": "Single-choice requires exactly one correct answer text."
                            },
                            status=400,
                        )
                    if not is_single and len(correct_texts) == 0:
                        return JsonResponse(
                            {
                                "error": "Multiple-choice requires at least one correct answer text."
                            },
                            status=400,
                        )
                    if len(options_texts) < 2:
                        return JsonResponse(
                            {"error": "MCQ requires at least two options."}, status=400
                        )

                    # Rebuild options, preserving ID/media if text matches, generating new if not
                    new_options = []
                    correct_ids = []
                    correct_text_set = set(correct_texts)
                    used_existing_ids = set()

                    for opt_text in options_texts:
                        if not isinstance(opt_text, str):
                            continue
                        existing_opt = existing_options_map_by_text.get(opt_text)
                        opt_id = None
                        media_file = None

                        if existing_opt and existing_opt["id"] not in used_existing_ids:
                            opt_id = existing_opt["id"]
                            media_file = existing_opt.get("media_filename")
                            used_existing_ids.add(opt_id)
                        else:  # New option text or ID already used, generate new ID
                            opt_id = str(uuid.uuid4())

                        new_options.append(
                            {
                                "id": opt_id,
                                "text": opt_text,
                                "media_filename": media_file,
                            }
                        )
                        if opt_text in correct_text_set:
                            correct_ids.append(opt_id)

                    # Final validation on mapped IDs
                    if len(correct_ids) != len(correct_text_set):
                        return JsonResponse(
                            {
                                "error": "Correct answer text mapping failed during update."
                            },
                            status=400,
                        )
                    if is_single and len(correct_ids) != 1:
                        return JsonResponse(
                            {
                                "error": "Correct answer ID mapping failed for single choice update."
                            },
                            status=400,
                        )

                    updated_question["options"] = new_options
                    updated_question["correct_answer"] = correct_ids

            # Save updated question back
            questions[question_index] = updated_question
            data["questions"] = questions
            # TODO: Update quiz associations if quiz_ids changed
            save_data(data)
            print(
                f"DEBUG [PUT Q Detail]: Question {question_id_str} updated (JSON only)."
            )
            return JsonResponse(
                {
                    "message": "Question updated (text/config only).",
                    "question": updated_question,
                }
            )

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON payload."}, status=400)
        except Exception as e:
            print(f"ERROR [PUT Q Detail]: {e}")
            traceback.print_exc()
            return JsonResponse(
                {"error": f"Error updating question: {str(e)}"}, status=500
            )

    # --- DELETE Request Logic (Add File Cleanup) ---
    elif request.method == "DELETE":
        print(f"DEBUG [DELETE Q Detail]: Request for Q ID: {question_id_str}")
        try:
            data = load_data()
            questions = data.get("questions", [])
            question_to_delete = None
            question_index = -1
            for index, q in enumerate(questions):
                if str(q.get("id")) == question_id_str:
                    question_to_delete = q
                    question_index = index
                    break
            if question_to_delete is None:
                return JsonResponse({"error": "Question not found."}, status=404)

            # --- Delete Associated Media Files ---
            files_to_delete = []
            if question_to_delete.get("media_filename"):
                files_to_delete.append(question_to_delete["media_filename"])
            if isinstance(question_to_delete.get("options"), list):
                for opt in question_to_delete["options"]:
                    if isinstance(opt, dict) and opt.get("media_filename"):
                        files_to_delete.append(opt["media_filename"])

            print(
                f"DEBUG [DELETE Q Detail]: Attempting to delete media files: {files_to_delete}"
            )
            deleted_count = 0
            if MEDIA_ROOT_DIR:  # Check if path is valid
                for filename in set(
                    files_to_delete
                ):  # Use set to avoid deleting same file twice
                    if not filename:
                        continue  # Skip empty filenames
                    try:
                        file_path = MEDIA_ROOT_DIR / filename
                        if file_path.is_file():
                            file_path.unlink()
                            print(
                                f"DEBUG [DELETE Q Detail]: Deleted media file: {file_path}"
                            )
                            deleted_count += 1
                        else:
                            print(
                                f"WARNING [DELETE Q Detail]: Media file not found for deletion: {file_path}"
                            )
                    except Exception as e:
                        print(
                            f"ERROR [DELETE Q Detail]: Failed to delete media file {filename}: {e}"
                        )
            print(f"DEBUG [DELETE Q Detail]: Deleted {deleted_count} media files.")
            # --- End Delete Media ---

            # Remove question from list
            deleted_question_text = questions.pop(question_index).get("text", "N/A")
            data["questions"] = questions
            # TODO: Remove question ID from associated quizzes in data['quizzes']
            save_data(data)
            print(
                f"DEBUG [DELETE Q Detail]: Question {question_id_str} removed from data."
            )

            return JsonResponse(
                {
                    "message": f'Question "{deleted_question_text[:30]}..." (ID: {question_id_str}) deleted successfully.'
                }
            )

        except Exception as e:
            print(f"ERROR [DELETE Q Detail]: {e}")
            traceback.print_exc()
            return JsonResponse(
                {"error": f"Error deleting question: {str(e)}"}, status=500
            )
    else:  # Should not happen
        return JsonResponse({"error": "Method Not Allowed"}, status=405)


# --- End question_detail_api ---


@csrf_exempt
@api_teacher_required
@require_http_methods(["GET", "PUT", "DELETE"])
def quiz_detail_api(request, quiz_id):
    """
    API endpoint for retrieving (GET), updating (PUT), or deleting (DELETE)
    a specific quiz by its ID. Requires teacher authentication. Uses JSON storage.
    """
    # Ensure quiz_id is string for consistent comparison later if needed elsewhere
    quiz_id_str = str(quiz_id)
    print(
        f"DEBUG [Detail API]: Request for Quiz ID: {quiz_id_str}, Method: {request.method}"
    )

    data = load_data()
    quizzes = data.get("quizzes", [])
    quiz = None
    quiz_index = -1

    # Find the quiz by ID
    for index, q in enumerate(quizzes):
        # Compare as strings for safety, even though URL converter gives UUID object
        if str(q.get("id")) == quiz_id_str:
            quiz = q
            quiz_index = index
            break

    if quiz is None:
        print(f"DEBUG [Detail API]: Quiz ID {quiz_id_str} not found.")
        return JsonResponse(
            {"error": f"Quiz with ID {quiz_id_str} not found."}, status=404
        )

    # --- GET Request ---
    if request.method == "GET":
        # Prepare the quiz data for response, ensuring no sets
        prepared_quiz = copy.deepcopy(quiz)  # Work on a deep copy

        # Convert potential sets to lists before returning JSON
        if isinstance(prepared_quiz.get("questions"), set):
            print(
                f"DEBUG [GET Detail]: Converting 'questions' set to list for quiz {quiz_id_str}"
            )
            prepared_quiz["questions"] = list(prepared_quiz["questions"])

        if isinstance(prepared_quiz.get("config"), dict):
            for k, v in prepared_quiz["config"].items():
                if isinstance(v, set):
                    print(
                        f"DEBUG [GET Detail]: Converting config key '{k}' set to list for quiz {quiz_id_str}"
                    )
                    prepared_quiz["config"][k] = list(v)
        # Add checks for other fields if necessary

        print(f"DEBUG [GET Detail]: Returning quiz detail for ID: {quiz_id_str}")
        return JsonResponse({"quiz": prepared_quiz})  # Return the cleaned quiz object

    # --- PUT Request (Update) ---
    elif request.method == "PUT":
        print(f"DEBUG [PUT Detail]: Processing update for quiz ID: {quiz_id_str}")
        try:
            request_data = json.loads(request.body)
            print(f"DEBUG [PUT Detail]: Received payload: {request_data}")

            # Work on a copy of the quiz data found
            updated_quiz = copy.deepcopy(quiz)

            # Update allowed fields (only update fields present in request_data)
            if "title" in request_data:
                updated_quiz["title"] = request_data["title"]
            if "description" in request_data:
                # Use get for optional field description, provide original as default
                updated_quiz["description"] = request_data.get(
                    "description", quiz.get("description", "")
                )

            # --- Update Configuration Block ---
            if "config" in request_data:
                print("DEBUG [PUT Detail]: Updating config block...")
                new_config_data = request_data["config"]
                if not isinstance(new_config_data, dict):
                    print(
                        f"ERROR [PUT Detail]: Invalid config data type received: {type(new_config_data)}"
                    )
                    return JsonResponse(
                        {
                            "error": "Invalid format for config data (must be an object)."
                        },
                        status=400,
                    )

                # --- Defensive handling for current_config ---
                current_config_raw = updated_quiz.get(
                    "config"
                )  # Get raw value from current quiz data
                if isinstance(current_config_raw, dict):
                    current_config = copy.deepcopy(
                        current_config_raw
                    )  # Work on a copy if it's already a dict
                    print("DEBUG [PUT Detail]: Copied existing config.")
                else:
                    # If it's missing, None, or not a dict, start fresh
                    print(
                        f"DEBUG [PUT Detail]: Existing config invalid or missing (type: {type(current_config_raw)}). Initializing empty config."
                    )
                    current_config = {}
                # --- current_config is now guaranteed to be a dictionary ---

                # Validate and update specific config fields from new_config_data
                # Duration
                if "duration" in new_config_data:
                    duration = new_config_data["duration"]
                    if duration is not None:
                        try:
                            duration = int(duration)
                            current_config["duration"] = (
                                duration if duration >= 0 else None
                            )
                        except (ValueError, TypeError):
                            current_config["duration"] = None  # Reset if invalid
                    else:
                        current_config["duration"] = None  # Allow setting to null

                # Pass Score
                if "pass_score" in new_config_data:
                    pass_score = new_config_data["pass_score"]
                    try:
                        pass_score = float(pass_score)
                        current_config["pass_score"] = (
                            pass_score if 0 <= pass_score <= 100 else 70
                        )
                    except (ValueError, TypeError):
                        current_config["pass_score"] = 70  # Default

                # Presentation Mode
                if "presentation_mode" in new_config_data:
                    presentation_mode = new_config_data["presentation_mode"]
                    current_config["presentation_mode"] = (
                        presentation_mode
                        if presentation_mode in ["all", "one-by-one"]
                        else "all"
                    )

                # Allow Back
                if "allow_back" in new_config_data:
                    current_config["allow_back"] = bool(new_config_data["allow_back"])

                # Randomize Questions
                if "randomize_questions" in new_config_data:
                    current_config["randomize_questions"] = bool(
                        new_config_data["randomize_questions"]
                    )

                # Shuffle Answers
                if "shuffle_answers" in new_config_data:
                    current_config["shuffle_answers"] = bool(
                        new_config_data["shuffle_answers"]
                    )

                # Assign the validated & updated config block back
                updated_quiz["config"] = current_config
                print(f"DEBUG [PUT Detail]: Updated config block: {current_config}")

            # --- Update Questions List ---
            if "questions" in request_data:
                print("DEBUG [PUT Detail]: Updating questions list...")
                # Basic validation: Ensure it's a list of strings (or can be converted)
                try:
                    # Ensure all items are strings (like UUIDs)
                    updated_questions_list = [
                        str(q_id) for q_id in request_data["questions"]
                    ]
                    updated_quiz["questions"] = updated_questions_list
                    print(
                        f"DEBUG [PUT Detail]: Set questions list to: {updated_questions_list}"
                    )
                except (TypeError, ValueError) as e:
                    print(
                        f"ERROR [PUT Detail]: Invalid questions list format: {request_data['questions']}. Error: {e}"
                    )
                    return JsonResponse(
                        {
                            "error": "Invalid format for questions list (must be a list of strings)."
                        },
                        status=400,
                    )

            # --- Update Archived Status ---
            if "archived" in request_data:
                updated_quiz["archived"] = bool(request_data["archived"])

            # Replace the old quiz dict with the updated one in the main list
            quizzes[quiz_index] = updated_quiz
            data["quizzes"] = quizzes
            save_data(data)  # Save changes back to JSON store (mocked in test)
            print(f"DEBUG [PUT Detail]: Quiz {quiz_id_str} saved successfully.")

            # --- Prepare response data (convert sets just in case) ---
            response_quiz = copy.deepcopy(updated_quiz)
            if isinstance(response_quiz.get("questions"), set):
                response_quiz["questions"] = list(response_quiz["questions"])
            if isinstance(response_quiz.get("config"), dict):
                for k, v in response_quiz["config"].items():
                    if isinstance(v, set):
                        response_quiz["config"][k] = list(v)

            return JsonResponse(
                {"message": "Quiz updated successfully.", "quiz": response_quiz}
            )

        except json.JSONDecodeError:
            print(
                f"ERROR [PUT Detail]: Invalid JSON format in request body for quiz {quiz_id_str}"
            )
            return JsonResponse(
                {"error": "Invalid JSON format in request body."}, status=400
            )
        except Exception as e:
            print(
                f"ERROR [PUT Detail]: Unexpected error updating quiz {quiz_id_str}: {e}"
            )
            import traceback

            traceback.print_exc()  # Print full traceback to server log
            return JsonResponse(
                {"error": f"An unexpected error occurred: {str(e)}"}, status=500
            )

    # --- DELETE Request ---
    elif request.method == "DELETE":
        print(f"DEBUG [DELETE Detail]: Processing delete for quiz ID: {quiz_id_str}")
        try:
            # Remove the quiz from the list
            deleted_quiz = quizzes.pop(quiz_index)  # Use pop on the original list
            deleted_quiz_title = deleted_quiz.get("title", "N/A")
            data["quizzes"] = quizzes  # Assign the modified list back
            save_data(data)  # Save changes
            print(f"DEBUG [DELETE Detail]: Quiz {quiz_id_str} deleted successfully.")

            return JsonResponse(
                {
                    "message": f'Quiz "{deleted_quiz_title}" (ID: {quiz_id_str}) deleted successfully.'
                }
            )
        except Exception as e:
            print(
                f"ERROR [DELETE Detail]: Unexpected error deleting quiz {quiz_id_str}: {e}"
            )
            import traceback

            traceback.print_exc()
            return JsonResponse(
                {"error": f"An unexpected error occurred: {str(e)}"}, status=500
            )


@api_teacher_required
@require_http_methods(["GET"])  # Export is typically a GET request
def quiz_export_api(request, quiz_id):
    """
    API endpoint to export a single quiz and its associated questions as JSON.
    """
    try:
        data = load_data()
        quizzes = data.get("quizzes", [])
        all_questions = data.get("questions", [])
        quiz_to_export = None

        # Find the quiz
        for q in quizzes:
            if str(q.get("id")) == str(quiz_id):  # Compare as strings
                quiz_to_export = q
                break

        if quiz_to_export is None:
            return JsonResponse(
                {"error": f"Quiz with ID {quiz_id} not found."}, status=404
            )

        # Find associated questions
        question_ids_in_quiz = set(quiz_to_export.get("questions", []))
        associated_questions = [
            q for q in all_questions if str(q.get("id")) in question_ids_in_quiz
        ]

        # Structure the export data
        export_data = {"quiz": quiz_to_export, "questions": associated_questions}

        # Prepare the file response
        json_data = json.dumps(export_data, indent=2)  # Pretty print JSON
        # Create a safe filename (e.g., replace spaces)
        filename_title = quiz_to_export.get("title", "quiz").replace(" ", "_")
        filename = f"quiz_{filename_title}_{quiz_id}.json"

        response = HttpResponse(json_data, content_type="application/json")
        response["Content-Disposition"] = (
            f'attachment; filename="{filename}"'  # Force download
        )
        response["Access-Control-Expose-Headers"] = (
            "Content-Disposition"  # Important for JS if needed later
        )

        print(f"DEBUG: Exporting quiz ID {quiz_id} as {filename}")  # Logging
        return response

    except Exception as e:
        print(f"Error exporting quiz {quiz_id}: {e}")
        import traceback

        traceback.print_exc()
        return JsonResponse(
            {"error": f"An unexpected error occurred during export: {str(e)}"},
            status=500,
        )


@csrf_exempt  # Use proper CSRF handling if form submitted directly later
@api_teacher_required
@require_http_methods(["POST"])
def quiz_import_api(request):
    """
    API endpoint to import a quiz and its questions from an uploaded JSON file.
    Generates new IDs for the imported items.
    """
    try:
        # Check if file is present
        if "quizFile" not in request.FILES:
            return JsonResponse({"error": "No quiz file provided."}, status=400)

        uploaded_file = request.FILES["quizFile"]

        # Basic validation: Check file extension (allow .json)
        if not uploaded_file.name.lower().endswith(".json"):
            return JsonResponse(
                {"error": "Invalid file type. Only .json files are accepted."},
                status=400,
            )

        # Read and parse JSON content
        try:
            import_data = json.load(uploaded_file)  # Reads directly from file stream
        except json.JSONDecodeError:
            return JsonResponse(
                {"error": "Invalid JSON format in the uploaded file."}, status=400
            )
        except Exception as e:
            return JsonResponse(
                {"error": f"Error reading file content: {str(e)}"}, status=400
            )

        # --- Validate and Process Import Data ---
        imported_quiz_data = import_data.get("quiz")
        imported_questions_data = import_data.get("questions", [])

        if not imported_quiz_data or not isinstance(imported_quiz_data, dict):
            return JsonResponse(
                {"error": 'Invalid JSON structure: Missing or invalid "quiz" object.'},
                status=400,
            )
        if not isinstance(imported_questions_data, list):
            return JsonResponse(
                {"error": 'Invalid JSON structure: "questions" must be a list.'},
                status=400,
            )

        # Load existing data
        data = load_data()
        all_quizzes = data.get("quizzes", [])
        all_questions = data.get("questions", [])

        # --- Process Questions First ---
        newly_created_question_ids = []
        old_to_new_question_id_map = {}
        for q_data in imported_questions_data:
            if (
                not isinstance(q_data, dict)
                or not q_data.get("text")
                or not q_data.get("type")
            ):
                print(f"Skipping invalid question data: {q_data}")  # Log skip
                continue  # Skip invalid question entries

            old_q_id = q_data.get("id")  # Get original ID for mapping
            new_q = q_data.copy()  # Create a copy to modify

            # Generate NEW IDs
            new_q_id = str(uuid.uuid4())
            new_q["id"] = new_q_id
            if old_q_id:  # Store mapping if original ID existed
                old_to_new_question_id_map[old_q_id] = new_q_id

            # Generate NEW Option IDs if MCQ
            if new_q.get("type") == "MCQ" and isinstance(new_q.get("options"), list):
                new_options = []
                old_to_new_option_id_map = {}
                correct_answer_texts = (
                    []
                )  # Need to rebuild correct answers based on new IDs

                # Find correct texts first using old IDs
                old_correct_option_ids = set(new_q.get("correct_answer", []))
                original_options = new_q.get("options", [])
                for opt in original_options:
                    if (
                        isinstance(opt, dict)
                        and opt.get("id") in old_correct_option_ids
                    ):
                        correct_answer_texts.append(opt.get("text"))

                # Generate new options and map correct answers
                new_correct_answer_ids = []
                for opt in original_options:
                    if not isinstance(opt, dict) or "text" not in opt:
                        continue  # Skip invalid option
                    new_opt_id = str(uuid.uuid4())
                    new_options.append({"id": new_opt_id, "text": opt["text"]})
                    # If this option's text was marked correct, use the new ID
                    if opt.get("text") in correct_answer_texts:
                        new_correct_answer_ids.append(new_opt_id)

                new_q["options"] = new_options
                new_q["correct_answer"] = new_correct_answer_ids
            else:
                # Clear options/answers if not MCQ
                new_q["options"] = []
                new_q["correct_answer"] = []

            # Add processed question to main list
            all_questions.append(new_q)
            newly_created_question_ids.append(new_q_id)  # Keep track for the quiz

        # --- Process Quiz ---
        new_quiz = imported_quiz_data.copy()
        new_quiz_id = str(uuid.uuid4())
        new_quiz["id"] = new_quiz_id

        # --- Modify Title ---
        original_title = new_quiz.get("title", "Untitled Quiz")
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        new_quiz["title"] = f"{original_title}_{timestamp}_(Imported)"
        print(f"DEBUG [Import]: Setting new quiz title to: {new_quiz['title']}")
        # --- End Modify Title ---

        # Update quiz's question list with NEW question IDs
        # Map old IDs from imported quiz's 'questions' list to new IDs
        old_question_ids_in_quiz = imported_quiz_data.get("questions", [])
        new_question_ids_for_quiz = [
            old_to_new_question_id_map[old_id]
            for old_id in old_question_ids_in_quiz
            if old_id
            in old_to_new_question_id_map  # Only include if the question was valid and processed
        ]
        # If no 'questions' key in import, use all newly created ones
        if not old_question_ids_in_quiz and newly_created_question_ids:
            new_question_ids_for_quiz = newly_created_question_ids

        new_quiz["questions"] = new_question_ids_for_quiz

        # Add basic validation for title
        if not new_quiz.get("title"):
            new_quiz["title"] = f"Imported_Quiz_{timestamp}"

        # Reset potentially sensitive/runtime data? (e.g., versions)
        new_quiz["versions"] = []
        new_quiz["archived"] = False  # Import as active

        # Add processed quiz to main list
        all_quizzes.append(new_quiz)

        # --- Save Updated Data ---
        data["quizzes"] = all_quizzes
        data["questions"] = all_questions
        save_data(data)

        print(
            f"DEBUG: Imported quiz '{new_quiz['title']}' (New ID: {new_quiz_id}) with {len(new_question_ids_for_quiz)} questions."
        )  # Logging
        return JsonResponse(
            {
                "message": "Quiz imported successfully!",
                "new_quiz_id": new_quiz_id,
                "new_quiz_title": new_quiz["title"],
                "questions_imported_count": len(newly_created_question_ids),
            },
            status=201,
        )

    except Exception as e:
        print(f"Error importing quiz: {e}")
        import traceback

        traceback.print_exc()
        return JsonResponse(
            {"error": f"An unexpected error occurred during import: {str(e)}"},
            status=500,
        )


@csrf_exempt  # If called via JS POST potentially, otherwise maybe GET
@require_http_methods(
    ["POST"]
)  # Let's use POST to receive the key potentially alongside student info later
def quiz_access_api(request):
    """
    API endpoint for students to access a quiz using a key.
    Validates the key, fetches quiz and question data, applies randomization,
    and returns the prepared quiz structure for the student to take.
    """
    try:
        request_data = json.loads(request.body)
        quiz_key = request_data.get("quiz_key")

        if not quiz_key:
            return JsonResponse({"error": "Quiz key is required."}, status=400)

        print(f"DEBUG: Quiz access requested with key: {quiz_key}")  # Log

        data = load_data()
        all_quizzes = data.get("quizzes", [])
        all_questions = data.get("questions", [])
        target_quiz = None
        quiz_metadata = None

        # --- Find Quiz by Access Key ---
        for quiz in all_quizzes:
            # Case-insensitive key matching recommended
            if (
                quiz.get("access_key")
                and quiz.get("access_key").upper() == quiz_key.upper()
            ):
                # Found the quiz, check if it's active
                if quiz.get("archived", False):
                    print(f"DEBUG: Quiz found for key {quiz_key} but is archived.")
                    return JsonResponse(
                        {"error": "This quiz is currently inactive."}, status=403
                    )  # Forbidden
                target_quiz = quiz
                break  # Stop searching once found

        if target_quiz is None:
            print(f"DEBUG: Invalid quiz key provided: {quiz_key}")
            return JsonResponse({"error": "Invalid quiz key."}, status=404)  # Not Found

        print(
            f"DEBUG: Found active quiz '{target_quiz.get('title')}' (ID: {target_quiz.get('id')}) for key {quiz_key}"
        )

        # --- Fetch Associated Questions ---
        question_ids_in_quiz = set(target_quiz.get("questions", []))
        quiz_questions = [
            q for q in all_questions if str(q.get("id")) in question_ids_in_quiz
        ]

        if not quiz_questions:
            print(f"DEBUG: Quiz {target_quiz.get('id')} has no associated questions.")
            return JsonResponse(
                {"error": "This quiz contains no questions."}, status=400
            )  # Bad Request

        print(
            f"DEBUG: Found {len(quiz_questions)} questions for quiz {target_quiz.get('id')}"
        )

        # --- Apply Randomization (Based on TCHR-6 config) ---
        quiz_config = target_quiz.get("config", {})

        # 1. Randomize Question Order
        if quiz_config.get("randomize_questions", False):
            print("DEBUG: Randomizing question order.")
            random.shuffle(quiz_questions)

        # 2. Shuffle Answer Options (for MCQs)
        if quiz_config.get("shuffle_answers", False):
            print("DEBUG: Shuffling MCQ answer options.")
            for question in quiz_questions:
                if question.get("type") == "MCQ" and isinstance(
                    question.get("options"), list
                ):
                    # Important: Shuffle a *copy* if you don't want to affect the stored data
                    # For now, let's shuffle in place just for this student's attempt
                    random.shuffle(question["options"])
                    # NOTE: Shuffling options means the stored `correct_answer` (which holds option IDs)
                    # is now potentially incorrect *relative to the shuffled order*.
                    # The grading logic (API-4) MUST compare student's selected option ID(s)
                    # against the original `correct_answer` list, NOT based on the shuffled index.

        # --- Prepare Data Payload for Student ---
        # Remove sensitive/internal data before sending
        prepared_quiz = {
            "id": target_quiz.get("id"),
            "title": target_quiz.get("title"),
            "description": target_quiz.get("description"),
            "config": {  # Send only relevant config for student UI
                "duration": quiz_config.get("duration"),
                "presentation_mode": quiz_config.get("presentation_mode", "all"),
                "allow_back": quiz_config.get("allow_back", True),
            },
            "questions": [],  # Prepare questions below
        }

        for q in quiz_questions:
            prepared_q = {
                "id": q.get("id"),
                "text": q.get("text"),
                "type": q.get("type"),
                "score": q.get("score", 1),
                "media_filename": q.get(
                    "media_filename"
                ),  # <<< Include question media filename
                "mcq_is_single_choice": q.get(
                    "mcq_is_single_choice", False
                ),  # <<< Include flag
                "options": [],
            }
            if q.get("type") == "MCQ" and isinstance(q.get("options"), list):
                prepared_q["options"] = [
                    {
                        "id": opt.get("id"),
                        "text": opt.get("text"),
                        "media_filename": opt.get(
                            "media_filename"
                        ),  # <<< Include option media filename
                    }
                    for opt in q.get("options", [])
                    if isinstance(opt, dict)
                ]
            prepared_quiz["questions"].append(prepared_q)

        print(
            f"DEBUG: Returning prepared quiz data for student. Quiz ID: {prepared_quiz['id']}, Question Count: {len(prepared_quiz['questions'])}"
        )
        return JsonResponse(prepared_quiz)

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid request format."}, status=400)
    except Exception as e:
        print(f"Error accessing quiz with key: {e}")
        import traceback

        traceback.print_exc()
        return JsonResponse(
            {"error": f"An unexpected error occurred: {str(e)}"}, status=500
        )


@csrf_exempt  # Student submissions likely won't have CSRF from standard forms
@require_http_methods(["POST"])
def quiz_submit_api(request, quiz_id):
    """
    API endpoint for students to submit their completed quiz answers.
    Performs grading, calculates score, stores attempt, and returns feedback.
    """
    try:
        submission_data = json.loads(request.body)
        print(f"DEBUG: Received submission for quiz_id: {quiz_id}")  # Log

        # --- Extract data from submission ---
        student_info_input = submission_data.get(
            "student_info"
        )  # <<< Use a temporary input variable name
        student_answers = submission_data.get("answers")
        start_time_iso = (
            submission_data.get("start_time")
            or datetime.datetime.now(datetime.timezone.utc).isoformat()
        )
        end_time_iso = (
            submission_data.get("end_time")
            or datetime.datetime.now(datetime.timezone.utc).isoformat()
        )
        submitted_due_to_timeout = submission_data.get(
            "submitted_due_to_timeout", False
        )

        # --- Validate and Process Student Info ---
        is_valid_student_data = False
        if isinstance(student_info_input, list) and len(student_info_input) > 0:
            # Check if all items in list are dicts with at least a name
            if all(
                isinstance(s, dict) and s.get("name", "").strip()
                for s in student_info_input
            ):  # Ensure name isn't empty string
                processed_student_list = student_info_input  # Assign the list directly
                is_valid_student_data = True
        elif (
            isinstance(student_info_input, dict)
            and student_info_input.get("name", "").strip()
        ):
            # Allow single dict, convert to list
            processed_student_list = [
                student_info_input
            ]  # Convert single student to list
            is_valid_student_data = True

        if not is_valid_student_data:
            return JsonResponse(
                {"error": "Missing or invalid student information (Name is required)."},
                status=400,
            )
        # --- End Student Info Validation ---

        # --- MODIFIED Answers Validation ---
        # Require 'answers' key to exist and be a dictionary, but allow it to be empty.
        # The grading loop will handle missing entries for individual questions.
        if student_answers is None or not isinstance(student_answers, dict):
            print(
                f"ERROR [Submit API]: Missing or invalid 'answers' field in payload. Type: {type(student_answers)}"
            )
            return JsonResponse(
                {"error": "Missing or invalid answers data format."}, status=400
            )

        # if not student_answers or not isinstance(student_answers, dict):
        #         return JsonResponse({'error': 'Missing or invalid answers data.'}, status=400)

        # --- Load Quiz and Question Data ---
        data = load_data()
        all_quizzes = data.get("quizzes", [])
        all_questions = data.get("questions", [])
        target_quiz = None
        for q in all_quizzes:
            if str(q.get("id")) == str(quiz_id):
                target_quiz = q
                break
        if target_quiz is None:
            return JsonResponse({"error": "Quiz not found."}, status=404)
        if target_quiz.get("archived"):
            return JsonResponse(
                {"error": "Cannot submit to an archived quiz."}, status=403
            )

        quiz_config = target_quiz.get("config", {})
        pass_score_threshold = Decimal(quiz_config.get("pass_score", 70))

        question_ids_in_quiz = set(target_quiz.get("questions", []))
        questions_in_quiz_dict = {
            str(q.get("id")): q
            for q in all_questions
            if str(q.get("id")) in question_ids_in_quiz
        }

        # --- Perform Grading ---
        total_score_achieved = Decimal(0)
        max_possible_score = Decimal(0)
        graded_details = []  # Optional detailed results

        for q_id, question_data in questions_in_quiz_dict.items():
            q_id_str = str(q_id)  # Ensure consistency
            question_score = Decimal(question_data.get("score", 1))  # Use Decimal
            max_possible_score += (
                question_score  # Add to max score for this quiz attempt
            )
            question_type = question_data.get("type")
            student_answer = student_answers.get(
                q_id_str
            )  # Get student's answer for this question

            is_correct = None  # None = not auto-graded/applicable, True/False otherwise
            score_awarded = Decimal(0)
            needs_manual_review = False
            # ... (MCQ / SHORT_TEXT grading) ...
            # total_score_achieved += score_awarded
            # graded_details.append({ # ... graded details ... })

            if student_answer is not None:  # Only grade if student provided an answer
                if question_type == "MCQ":
                    correct_option_ids = set(question_data.get("correct_answer", []))
                    is_single_choice_q = question_data.get(
                        "mcq_is_single_choice", False
                    )  # <<< Get flag

                    if isinstance(student_answer, list):
                        selected_option_ids = set(student_answer)

                        # --- Modified Correctness Check ---
                        if is_single_choice_q:
                            # Single choice: correct if exactly one selected AND it's the correct one
                            is_correct = len(
                                selected_option_ids
                            ) == 1 and selected_option_ids.issubset(correct_option_ids)
                            # We assume correct_option_ids also has only 1 item if is_single_choice_q is true due to validation
                        else:
                            # Multiple choice: exact match required
                            is_correct = selected_option_ids == correct_option_ids
                        # --- End Modified Check ---

                        if is_correct:
                            score_awarded = question_score
                    else:
                        is_correct = False  # Invalid answer format

                elif question_type == "SHORT_TEXT":
                    review_mode = question_data.get(
                        "short_answer_review_mode", "manual"
                    )
                    if review_mode == "auto":
                        correct_text = question_data.get("short_answer_correct_text")
                        if correct_text is not None and isinstance(student_answer, str):
                            # Case-INSENSITIVE comparison after trimming whitespace
                            is_correct = (
                                student_answer.strip().lower() == correct_text.lower()
                            )
                            if is_correct:
                                score_awarded = question_score
                        else:
                            is_correct = False  # Cannot auto-grade if no correct text stored or invalid student answer
                    else:  # Manual review mode
                        needs_manual_review = True
                        is_correct = None  # Mark as needing review

                # Add logic for other question types here later

            # Accumulate total score
            total_score_achieved += score_awarded

            # Store detailed result (optional)
            graded_details.append(
                {
                    "question_id": q_id_str,
                    "is_correct": is_correct,
                    "score_awarded": float(
                        score_awarded.quantize(Decimal("0.01"), ROUND_HALF_UP)
                    ),  # Store as float
                    "needs_manual_review": needs_manual_review,
                }
            )

        # Calculate percentage
        percentage = Decimal(0)
        if max_possible_score > 0:
            percentage = (total_score_achieved / max_possible_score) * 100
            # Round to reasonable precision, e.g., 2 decimal places
            percentage = percentage.quantize(Decimal("0.01"), ROUND_HALF_UP)

        passed = percentage >= pass_score_threshold

        # --- Prepare Attempt Record ---
        attempt_id = str(uuid.uuid4())
        # Convert timestamps if they were milliseconds
        # start_time_iso = datetime.datetime.fromtimestamp(start_time_ms / 1000).isoformat() if start_time_ms else None
        # end_time_iso = datetime.datetime.fromtimestamp(end_time_ms / 1000).isoformat() if end_time_ms else datetime.datetime.now().isoformat()
        # Using simple now() for timestamps for now, as frontend isn't sending them yet fully
        start_time_iso = (
            submission_data.get("start_time")
            or datetime.datetime.now(datetime.timezone.utc).isoformat()
        )
        end_time_iso = (
            submission_data.get("end_time")
            or datetime.datetime.now(datetime.timezone.utc).isoformat()
        )
        submitted_due_to_timeout = submission_data.get(
            "submitted_due_to_timeout", False
        )

        new_attempt = {
            "attempt_id": attempt_id,
            "quiz_id": str(quiz_id),
            "quiz_title_at_submission": target_quiz.get("title", "N/A"),
            "students": processed_student_list,  # <<< Use the reliably assigned list variable
            "answers": student_answers,
            "score_achieved": float(
                total_score_achieved.quantize(Decimal("0.01"), ROUND_HALF_UP)
            ),
            "max_possible_score": float(
                max_possible_score.quantize(Decimal("0.01"), ROUND_HALF_UP)
            ),
            "percentage": float(percentage),
            "passed": bool(passed),
            "pass_score_threshold": float(pass_score_threshold),
            "start_time": start_time_iso,
            "end_time": end_time_iso,
            "submitted_due_to_timeout": submitted_due_to_timeout,
            "graded_details": graded_details,
        }

        # --- Save Attempt ---
        all_attempts = data.get("attempts", [])
        all_attempts.append(new_attempt)
        data["attempts"] = all_attempts
        save_data(data)
        print(
            f"DEBUG: Stored attempt {attempt_id} for quiz {quiz_id}. Score: {percentage}%"
        )

        # --- Return Feedback ---
        feedback_payload = {
            "attempt_id": attempt_id,
            "score": float(percentage),
            "passed": bool(passed),
            "max_score": float(max_possible_score),
            "achieved_score": float(total_score_achieved),
        }
        return JsonResponse(feedback_payload, status=201)

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid submission format."}, status=400)
    except Exception as e:
        print(f"Error processing submission for quiz {quiz_id}: {e}")
        import traceback

        traceback.print_exc()
        # Return the specific Python error message to help debugging
        return JsonResponse(
            {"error": f"An unexpected error occurred during submission: {str(e)}"},
            status=500,
        )


@api_teacher_required
@require_http_methods(["GET"])
def get_quiz_attempts_api(request, quiz_id):
    """
    API endpoint for teachers to retrieve a summary list of attempts
    for a specific quiz.
    """
    print(f"DEBUG [Attempts API]: Fetching attempts for quiz {quiz_id}")
    try:
        data = load_data()
        all_attempts = data.get("attempts", [])
        quiz_id_str = str(quiz_id)

        # Filter attempts for the specific quiz
        quiz_attempts = [
            attempt
            for attempt in all_attempts
            if str(attempt.get("quiz_id")) == quiz_id_str
        ]
        print(
            f"DEBUG [Attempts API]: Found {len(quiz_attempts)} raw attempts for quiz."
        )

        # Sort attempts, e.g., by submission time descending
        quiz_attempts.sort(key=lambda x: x.get("end_time", ""), reverse=True)

        # Prepare summarized data to return
        summarized_attempts = []
        for att in quiz_attempts:
            # --- CORRECTED STUDENT DETAIL EXTRACTION ---
            students_in_attempt = att.get(
                "students", []
            )  # Get the list of students for this attempt
            student_name_display = "N/A"
            student_class_display = ""
            student_id_display = ""

            if isinstance(students_in_attempt, list) and len(students_in_attempt) > 0:
                # Format names (handle single or multiple)
                student_name_display = " & ".join(
                    [
                        str(
                            s.get("name", "N/A")
                        ).strip()  # Ensure name is string and trim
                        for s in students_in_attempt
                        if isinstance(s, dict)  # Check item is dict
                    ]
                )
                # Get class/id from the FIRST student in the list for summary display
                first_student = students_in_attempt[0]
                if isinstance(first_student, dict):
                    student_class_display = str(first_student.get("class", "")).strip()
                    student_id_display = str(first_student.get("id", "")).strip()

            elif isinstance(
                att.get("student_info"), dict
            ):  # Fallback for old structure
                print(
                    f"WARNING [Attempts API]: Attempt {att.get('attempt_id')} using old 'student_info'."
                )
                legacy_info = att["student_info"]
                student_name_display = str(legacy_info.get("name", "N/A")).strip()
                student_class_display = str(legacy_info.get("class", "")).strip()
                student_id_display = str(legacy_info.get("id", "")).strip()
            # --- END CORRECTION ---

            # Prepare the summary dictionary for this attempt
            summary = {
                "attempt_id": att.get("attempt_id"),
                "student_name": student_name_display,  # Use extracted value
                "student_class": student_class_display,  # Use extracted value
                "student_id_number": student_id_display,  # Use extracted value
                "score_percentage": att.get("percentage"),
                "passed": att.get("passed"),
                "submission_time": att.get("end_time"),
                "submitted_due_to_timeout": att.get("submitted_due_to_timeout", False),
            }
            summarized_attempts.append(summary)

        print(
            f"DEBUG [Attempts API]: Returning {len(summarized_attempts)} summarized attempts."
        )
        return JsonResponse({"attempts": summarized_attempts})

    except Exception as e:
        print(f"Error fetching attempts for quiz {quiz_id}: {e}")
        traceback.print_exc()
        return JsonResponse(
            {"error": f"An unexpected error occurred: {str(e)}"}, status=500
        )


def _get_attempts_for_quiz(quiz_id_to_find):
    """Helper function to fetch and sort attempts for reuse."""
    data = load_data()
    all_attempts = data.get("attempts", [])
    quiz_id_str = str(quiz_id_to_find)
    quiz_attempts = [
        attempt
        for attempt in all_attempts
        if str(attempt.get("quiz_id")) == quiz_id_str
    ]
    quiz_attempts.sort(key=lambda x: x.get("end_time", ""), reverse=True)
    # Also fetch quiz title for filename
    quiz_title = "UnknownQuiz"
    all_quizzes = data.get("quizzes", [])
    for q in all_quizzes:
        if str(q.get("id")) == quiz_id_str:
            quiz_title = q.get("title", "Quiz").replace(
                " ", "_"
            )  # Safe title for filename
            break
    return quiz_attempts, quiz_title


@api_teacher_required
@require_http_methods(["GET"])
def export_quiz_attempts_json_api(request, quiz_id):
    """Exports attempts for a specific quiz as a JSON file."""
    try:
        quiz_attempts, quiz_title = _get_attempts_for_quiz(quiz_id)
        if not quiz_attempts:
            # Optionally return empty file or error? Let's return empty for now.
            pass

        # Select fields for export (can be more detailed than summary list)
        export_data = [
            {
                "Attempt ID": att.get("attempt_id"),
                "Student Name": att.get("student_info", {}).get("name", "N/A"),
                "Student Class": att.get("student_info", {}).get("class", ""),
                "Student ID/Number": att.get("student_info", {}).get("id", ""),
                "Score (%)": att.get("percentage"),
                "Score Achieved": att.get("score_achieved"),
                "Max Score": att.get("max_possible_score"),
                "Passed": att.get("passed"),
                "Submission Time (UTC)": att.get("end_time"),
                "Timed Out": att.get("submitted_due_to_timeout", False),
                "Answers": att.get("answers"),  # Include raw answers in JSON export
                "Graded Details": att.get("graded_details"),  # Include grading details
            }
            for att in quiz_attempts
        ]

        json_export_data = json.dumps({"quiz_results": export_data}, indent=2)
        filename = f"results_{quiz_title}_{quiz_id}.json"
        response = HttpResponse(json_export_data, content_type="application/json")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        # Handle errors appropriately, maybe return JSON error response
        print(f"Error exporting JSON results for quiz {quiz_id}: {e}")
        return JsonResponse({"error": "Failed to generate JSON export."}, status=500)


@api_teacher_required
@require_http_methods(["GET"])
def export_quiz_attempts_excel_api(request, quiz_id):
    """Exports attempts for a specific quiz as an Excel (.xlsx) file."""
    try:
        quiz_attempts, quiz_title = _get_attempts_for_quiz(quiz_id)

        # Create Excel workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Quiz Results ({quiz_title[:20]})"  # Sheet title limit

        # Define headers
        headers = [
            "Attempt ID",
            "Student Name",
            "Student Class",
            "Student ID/Number",
            "Score (%)",
            "Passed",
            "Submission Time (UTC)",
            "Timed Out",
            "Score Achieved",
            "Max Score",
            # Avoid exporting raw answers/details to basic Excel for simplicity,
            # can be added as separate sheets or complex cells if needed later.
        ]
        ws.append(headers)

        # Add data rows
        for att in quiz_attempts:
            passed_text = "Yes" if att.get("passed") else "No"
            timeout_text = "Yes" if att.get("submitted_due_to_timeout", False) else "No"
            row = [
                att.get("attempt_id", ""),
                att.get("student_info", {}).get("name", "N/A"),
                att.get("student_info", {}).get("class", ""),
                att.get("student_info", {}).get("id", ""),
                att.get("percentage", ""),
                passed_text,
                att.get("end_time", ""),
                timeout_text,
                att.get("score_achieved", ""),
                att.get("max_possible_score", ""),
            ]
            ws.append(row)

        # Optional: Adjust column widths
        for i, column_cells in enumerate(ws.columns):
            max_length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = max_length + 2
            ws.column_dimensions[get_column_letter(i + 1)].width = adjusted_width

        # Prepare HTTP response
        filename = f"results_{quiz_title}_{quiz_id}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)  # Save workbook content to the response
        return response

    except Exception as e:
        print(f"Error exporting Excel results for quiz {quiz_id}: {e}")
        return JsonResponse({"error": "Failed to generate Excel export."}, status=500)


# Helper function (can be moved to utils if desired)
def generate_quiz_key(length=6):
    characters = string.ascii_uppercase + string.digits
    # Ensure reasonable uniqueness? Simple random for now.
    return "".join(random.choice(characters) for i in range(length))


@api_teacher_required
@require_http_methods(["POST"])  # Use POST for action that modifies data
def quiz_regenerate_key_api(request, quiz_id):
    """
    API endpoint for regenerating the access key for a specific quiz.
    """
    try:
        data = load_data()
        quizzes = data.get("quizzes", [])
        quiz_to_update = None
        quiz_index = -1

        # Find the quiz
        for index, q in enumerate(quizzes):
            if str(q.get("id")) == str(quiz_id):
                quiz_to_update = q
                quiz_index = index
                break

        if quiz_to_update is None:
            return JsonResponse(
                {"error": f"Quiz with ID {quiz_id} not found."}, status=404
            )

        # Generate a new key
        new_key = generate_quiz_key()
        # TODO: Optional - Add check here to ensure new_key isn't already used by another *active* quiz

        print(
            f"DEBUG [Regen Key]: Regenerating key for quiz {quiz_id}. Old: {quiz_to_update.get('access_key')}, New: {new_key}"
        )

        # Update the quiz object
        quiz_to_update["access_key"] = new_key
        quizzes[quiz_index] = quiz_to_update  # Place updated quiz back in list
        data["quizzes"] = quizzes

        # Save the entire data structure back
        save_data(data)

        return JsonResponse(
            {"message": "Access key regenerated successfully.", "new_key": new_key}
        )

    except Exception as e:
        print(f"Error regenerating key for quiz {quiz_id}: {e}")
        import traceback

        traceback.print_exc()
        return JsonResponse(
            {"error": f"An unexpected error occurred: {str(e)}"}, status=500
        )
